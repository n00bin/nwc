#!/usr/bin/env python3
"""Apply edits from the Google-Sheets workbook back to the database.

  python scripts/sheet_import.py <downloaded.xlsx>            (dry run)
  python scripts/sheet_import.py <downloaded.xlsx> --apply    (write JSON)

Diff-based and conservative — the JSON in ../data/ is the source of truth:
  - Only cells that DIFFER from the current database are touched.
  - Each changed cell is parsed by the same encoding sheet_export.py used
    (lists, "Stat: value" maps, [JSON] columns) and validated; a cell that
    doesn't parse is REJECTED with a reason — everything else still applies.
  - Rows whose id isn't in the database are treated as NEW items (id
    assigned automatically). Rows missing from the sheet are NOT deleted —
    deletions stay a deliberate JSON-side act.
After --apply, run build-data.py and review `git diff ../data` like any
other data change.
"""
import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

DATA = Path(__file__).resolve().parent.parent.parent / "data"


def flatten_value(v):  # must mirror sheet_export.py exactly
    if v is None:
        return "", "scalar"
    if isinstance(v, (str, int, float, bool)):
        return v, "scalar"
    if isinstance(v, list):
        if all(isinstance(x, (str, int, float, bool)) for x in v):
            return ", ".join(str(x) for x in v), "list"
        return json.dumps(v, ensure_ascii=False, separators=(",", ":")), "json"
    if isinstance(v, dict):
        if all(isinstance(x, (str, int, float, bool)) for x in v.values()):
            return "\n".join(f"{k}: {x}" for k, x in v.items()), "map"
        return json.dumps(v, ensure_ascii=False, separators=(",", ":")), "json"
    return str(v), "scalar"


def cells_equal(cell, current):
    """Compare a sheet cell against the current JSON value, tolerating the
    type wobble Excel introduces (1560 vs '1560', 1.0 vs 1, trailing space)."""
    rendered, _ = flatten_value(current)
    if cell is None:
        cell = ""
    if isinstance(cell, str):
        cell = cell.replace("\r\n", "\n").strip()
    if isinstance(rendered, str):
        rendered = rendered.strip()
    if isinstance(cell, (int, float)) and isinstance(rendered, (int, float)):
        return float(cell) == float(rendered)
    return str(cell) == str(rendered)


def parse_number(s):
    if isinstance(s, (int, float)):
        return int(s) if float(s).is_integer() else float(s)
    t = str(s).strip().replace(",", "")
    f = float(t)
    return int(f) if f.is_integer() else f


def parse_cell(cell, kind, current):
    """Sheet cell -> JSON value, based on column kind + the current value's
    type. Raises ValueError with a human reason when malformed."""
    text = "" if cell is None else (cell if isinstance(cell, str) else cell)
    if kind == "json":
        s = str(text).strip()
        if s == "":
            return None
        try:
            return json.loads(s)
        except json.JSONDecodeError as e:
            raise ValueError(f"not valid JSON ({e.msg} at char {e.pos})")
    if kind == "list":
        s = str(text).strip()
        if s == "":
            return []
        parts = [p.strip() for p in s.split(",") if p.strip()]
        if current and all(isinstance(x, (int, float)) for x in current):
            return [parse_number(p) for p in parts]
        return parts
    if kind == "map":
        s = str(text).replace("\r\n", "\n").strip()
        if s == "":
            return {}
        out = {}
        for line in s.split("\n"):
            line = line.strip()
            if not line:
                continue
            m = re.match(r"^(.+?):\s*(.+)$", line)
            if not m:
                raise ValueError(f"line '{line}' is not 'Stat: value'")
            key, val = m.group(1).strip(), m.group(2).strip()
            try:
                out[key] = parse_number(val)
            except ValueError:
                out[key] = val
            if not key:
                raise ValueError(f"empty stat name in '{line}'")
        return out
    # scalar — steer by the current value's type when we have one
    if isinstance(text, str):
        text = text.strip()
    if text == "" or text is None:
        return None
    if isinstance(current, bool):
        s = str(text).strip().lower()
        if s in ("true", "yes", "1"): return True
        if s in ("false", "no", "0"): return False
        raise ValueError(f"expected true/false, got '{text}'")
    if isinstance(current, (int, float)):
        try:
            return parse_number(text)
        except ValueError:
            raise ValueError(f"expected a number, got '{text}'")
    if isinstance(text, (int, float)) and current is None:
        return parse_number(text)
    return str(text)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="workbook downloaded from Google Sheets")
    ap.add_argument("--apply", action="store_true", help="write changes to ../data/*.json")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    changes, rejects, new_items = [], [], []

    for ws in wb.worksheets:
        if ws.title == "READ ME":
            continue
        path = DATA / f"{ws.title}.json"
        if not path.exists():
            rejects.append(f"[{ws.title}] no matching data file — tab ignored")
            continue
        items = json.loads(path.read_text(encoding="utf-8"))
        by_id = {it.get("id"): it for it in items if it.get("id") is not None}
        by_name = {}
        for it in items:
            by_name.setdefault(str(it.get("name", "")).strip().lower(), it)

        header = [c.value for c in ws[1]]
        cols = []
        for h in header:
            if h is None:
                cols.append((None, None)); continue
            name = str(h)
            kind = "json" if name.endswith(" [JSON]") else None
            cols.append((name.replace(" [JSON]", ""), kind))

        dirty = False
        for row in ws.iter_rows(min_row=2):
            if all(c.value in (None, "") for c in row):
                continue
            vals = {cols[i][0]: row[i].value for i in range(min(len(cols), len(row))) if cols[i][0]}
            rid = vals.get("id")
            target = None
            if rid not in (None, ""):
                try:
                    target = by_id.get(int(rid))
                except (TypeError, ValueError):
                    target = None
            if target is None and vals.get("name"):
                target = by_name.get(str(vals["name"]).strip().lower())

            if target is None:
                # NEW item — build it from every non-empty cell
                fresh = {}
                ok = True
                for i in range(min(len(cols), len(row))):
                    col, forced = cols[i]
                    if not col or row[i].value in (None, ""):
                        continue
                    if col == "id":
                        continue
                    kind = forced or ("scalar")
                    try:
                        fresh[col] = parse_cell(row[i].value, kind, None)
                    except ValueError as e:
                        rejects.append(f"[{ws.title}] NEW '{vals.get('name','?')}' col {col}: {e}")
                        ok = False
                if ok and fresh.get("name"):
                    nid = max([it.get("id") or 0 for it in items] + [0]) + 1
                    fresh = {"id": nid, **fresh}
                    items.append(fresh)
                    new_items.append(f"[{ws.title}] + {fresh['name']} (id {nid})")
                    dirty = True
                elif ok:
                    rejects.append(f"[{ws.title}] row {row[0].row}: new row has no name — skipped")
                continue

            for i in range(min(len(cols), len(row))):
                col, forced = cols[i]
                if not col or col == "id":
                    continue
                current = target.get(col)
                if cells_equal(row[i].value, current):
                    continue
                _, cur_kind = flatten_value(current)
                kind = forced or (cur_kind if current is not None else "scalar")
                try:
                    new_val = parse_cell(row[i].value, kind, current)
                except ValueError as e:
                    rejects.append(f"[{ws.title}] {target.get('name','?')} (id {target.get('id')}) col {col}: {e}")
                    continue
                old_txt, _ = flatten_value(current)
                new_txt, _ = flatten_value(new_val)
                changes.append(f"[{ws.title}] {target.get('name','?')} (id {target.get('id')}) {col}: "
                               f"{str(old_txt)[:60]!r} -> {str(new_txt)[:60]!r}")
                if new_val is None and col in target:
                    del target[col]
                else:
                    target[col] = new_val
                dirty = True

        if dirty and args.apply:
            path.write_text(json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"=== {len(changes)} cell change(s), {len(new_items)} new item(s), {len(rejects)} rejected ===")
    for c in changes:
        print("  ~", c)
    for n in new_items:
        print("  +", n)
    if rejects:
        print("\nREJECTED (fix the cell and re-download, or do it JSON-side):")
        for r in rejects:
            print("  X", r)
    if not args.apply and (changes or new_items):
        print("\nDry run — nothing written. Re-run with --apply to commit these to ../data/, "
              "then run build-data.py.")
    elif args.apply and (changes or new_items):
        print("\nApplied. Next: python build-data.py  +  review `git diff` in ../data")


if __name__ == "__main__":
    sys.exit(main())
