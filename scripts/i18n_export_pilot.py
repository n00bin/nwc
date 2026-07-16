#!/usr/bin/env python3
"""
i18n pilot — export the English UI strings for the HOME PAGE + NAV + FOOTER
into a translator-friendly sheet (English filled, Russian blank).

Hand-curated from index.html and js/shared.js (renderNav/footer). Auto-scraping
the HTML would drag in emoji, decorative arrows and layout noise, so the strings
are listed explicitly here — this file doubles as the reviewed source list.

Design: the future i18n layer keys by the ENGLISH string itself
(t("Mounts") -> Russian, or falls back to "Mounts" if unfilled). So this sheet
needs no key column, dedupes automatically (one row per unique English string),
and a blank Russian cell simply leaves that string in English. Nothing breaks
if the sheet is half-finished.

Output: docs/i18n/ui_strings_pilot.csv  (canonical, git-diffable)
        docs/i18n/ui_strings_pilot.xlsx (if openpyxl present — matches the
        upload-to-Google-Sheets flow n00b already uses for data)

Run:  python3 scripts/i18n_export_pilot.py
"""
import os, csv, collections

HERE = os.path.dirname(os.path.abspath(__file__))
WEB = os.path.dirname(HERE)
OUTDIR = os.path.join(WEB, "docs", "i18n")
os.makedirs(OUTDIR, exist_ok=True)

# (english, where_it_appears, is_brand)
# is_brand = a proper noun / brand. Fine to leave blank (stays English) unless
# there's an official Russian form the player wants to use.
STRINGS = [
    # --- Header ---
    ("Neverwinter Compendium", "Home: big title at the top (also the footer copyright line)", True),
    ("Everything you need to gear up in Neverwinter — mounts, companions, artifacts, and more, all in one friendly place.",
     "Home: subtitle under the title", False),

    # --- Home / News tabs ---
    ("Home", "Tab at top of home page; also the first nav menu item", False),
    ("News", "Tab at top of home page (site news feed)", False),

    # --- Preview banner ---
    ("is live", "Home: preview banner, shown after the module name (e.g. \"Mod 33 Preview is live\")", False),
    ("Gear, comps, and screenshots from the upcoming module", "Home: preview banner subtitle", False),
    ("View Preview", "Home: preview banner button", False),

    # --- Landing card titles ---
    ("Mounts", "Home card title + nav menu item", False),
    ("Companions", "Home card title + nav menu item", False),
    ("Artifacts", "Home card title + nav menu item", False),
    ("Consumables", "Home card title + nav menu item", False),
    ("Mekaniks", "Home card title + nav menu item (stylized spelling of \"Mechanics\")", False),
    ("Campaign Boosters", "Home card title + nav menu item", False),
    ("Professions", "Home card title + nav menu item", False),
    ("NW Patch Notes", "Home card title + nav menu item", False),
    ("Reports", "Home card title + nav menu item", False),
    ("Mod 33 Preview", "Home card title (upcoming-module preview)", False),
    ("YouTube", "Home card title (link to the YouTube channel)", True),
    ("Join N00bin Network", "Home card title (YouTube membership)", True),

    # --- Landing card subtitles ---
    ("Stats, caps & formulas", "Home: subtitle on the Mekaniks card", False),
    ("Currency boost items & companions", "Home: subtitle on the Campaign Boosters card", False),
    ("Artisans, tips & masterwork", "Home: subtitle on the Professions card", False),
    ("Auto-updated daily from Arc Games", "Home: subtitle on the Patch Notes card", False),
    ("Bugs, missing items & suggestions", "Home: subtitle on the Reports card", False),
    ("Gear, comps & preview screenshots", "Home: subtitle on the Mod 33 Preview card", False),
    ("Watch N00bin on YouTube", "Home: subtitle on the YouTube card", False),
    ("Become a member of the community", "Home: subtitle on the Join card", False),

    # --- News view ---
    ("No news just yet — check back soon!", "News tab: shown when there are no news posts", False),
    ("Feature", "News post tag (blue) — marks a new feature", False),
    ("Fix", "News post tag (green) — marks a bug fix", False),
    ("Data", "News post tag (yellow) — marks new data", False),

    # --- Nav (only the items not already covered by a card title above) ---
    ("Creators & Tools", "Nav menu item (community creators + site tools)", False),
    ("The N00bin Network", "Nav + footer: link to the YouTube channel", True),
    ("Join on YouTube", "Nav + footer: YouTube membership link", False),

    # --- Footer ---
    ("Want to collaborate or contribute data? Reach out:", "Footer: invitation line before the email address", False),
    ("Browse the full item database", "Footer: link to the full database", False),
    ("every companion, mount, gear piece, artifact & more", "Footer: text after the database link", False),

    # --- Count labels (DYNAMIC: shown after a number, e.g. \"339 mounts\") ---
    # Russian plural after a number is tricky (1 скакун / 2 скакуна / 5 скакунов).
    # For the pilot, give the plural form used after big numbers (5+) — that's
    # what almost every count on the site is. We can refine later if it matters.
    ("Loading...", "Home cards: placeholder shown for a moment before counts load", False),
    ("mounts", "Home: the word after the number, e.g. \"339 mounts\" (plural, after 5+)", False),
    ("companions", "Home: the word after the number, e.g. \"268 companions\" (plural, after 5+)", False),
    ("buffs", "Home: the word after the number, e.g. \"99 buffs\" (plural, after 5+)", False),
    ("artifacts", "Home: the word after the number, e.g. \"140 artifacts\" (plural, after 5+)", False),
    ("sets", "Home: the word after the number, e.g. \"38 sets\" (plural, after 5+)", False),
]

# Dedupe by English string (keep first context; note if it recurs).
seen = collections.OrderedDict()
for eng, ctx, brand in STRINGS:
    if eng in seen:
        continue
    seen[eng] = (ctx, brand)

rows = [(eng, ctx, brand) for eng, (ctx, brand) in seen.items()]

HEADER = ["English (do not edit)", "Russian (fill this in)", "Where it appears", "Brand?"]

# --- CSV (canonical) ---
csv_path = os.path.join(OUTDIR, "ui_strings_pilot.csv")
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(HEADER)
    for eng, ctx, brand in rows:
        w.writerow([eng, "", ctx, "brand" if brand else ""])

# --- XLSX (optional, friendlier for Google Sheets) ---
xlsx_path = os.path.join(OUTDIR, "ui_strings_pilot.xlsx")
xlsx_made = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Home + Nav + Footer"

    head_fill = PatternFill("solid", fgColor="2F2F2F")
    head_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(HEADER, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    russ_fill = PatternFill("solid", fgColor="FFF7E6")  # tint the column he fills
    for r, (eng, ctx, brand) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=eng)
        rc = ws.cell(row=r, column=2, value="")
        rc.fill = russ_fill
        ws.cell(row=r, column=3, value=ctx)
        ws.cell(row=r, column=4, value="brand" if brand else "")

    widths = [58, 34, 60, 8]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    for r in range(1, len(rows) + 2):
        for c in (1, 3):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(xlsx_path)
    xlsx_made = True
except ImportError:
    pass

print("UI strings (unique): %d" % len(rows))
print("CSV : %s" % os.path.relpath(csv_path, WEB))
if xlsx_made:
    print("XLSX: %s" % os.path.relpath(xlsx_path, WEB))
else:
    print("XLSX: skipped (openpyxl not installed) — CSV imports into Google Sheets fine")
