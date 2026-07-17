#!/usr/bin/env python3
"""
i18n pilot — export the NAMES shown on the Mounts page for Russian translation.
These are the in-game proper names (mounts, mount powers, insignia bonuses,
insignias, collars) that are kept English via translate="no" so the browser
can't mangle them. When RU mode is on we'll show Dark Lord's Russian versions
instead (English stays the default, so console players are unaffected).

Source of the English names: ../data/*.json (verified game data — NOT invented).
Russian is filled in by Dark Lord from the official RU client.

Imported as a NEW TAB "Mounts - Names" in the existing translation doc
(File -> Import -> Insert new sheet(s)).

Output: docs/i18n/ui_strings_mount_names.csv / .xlsx
Run:    G:/Python/python.exe scripts/i18n_export_mount_names.py
"""
import os, csv, json, io

HERE = os.path.dirname(os.path.abspath(__file__))
WEB = os.path.dirname(HERE)
SRC = os.path.normpath(os.path.join(WEB, "..", "data"))
OUTDIR = os.path.join(WEB, "docs", "i18n")
os.makedirs(OUTDIR, exist_ok=True)

def load(f):
    return json.load(io.open(os.path.join(SRC, f), encoding="utf-8"))

def names_of(data):
    if isinstance(data, dict):
        data = data.get("insignias", [])
    out = []
    for x in data:
        if isinstance(x, dict):
            n = (x.get("name") or "").strip()
            if n:
                out.append(n)
    return out

# (english_names, category, note)
SECTIONS = [
    (names_of(load("mounts.json")),                 "Mount",          ""),
    (names_of(load("mount_combat_powers.json")),    "Combat Power",   ""),
    (names_of(load("mount_equip_powers.json")),     "Equip Power",    ""),
    (names_of(load("mount_insignia_bonuses.json")), "Insignia Bonus", ""),
    # Patterned / repetitive names — lower value, marked optional.
    (names_of(load("mount_insignias.json")),        "Insignia",       "optional"),
    (names_of(load("mount_collars.json")),          "Collar",         "optional"),
]

rows = []
seen = set()
for raw_names, cat, note in SECTIONS:
    for n in sorted(set(raw_names), key=str.lower):
        if n in seen:
            continue
        seen.add(n)
        rows.append((n, cat, note))

HEADER = ["English name (do not edit)", "Russian name (fill this in)", "Category", "Note"]

csv_path = os.path.join(OUTDIR, "ui_strings_mount_names.csv")
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(HEADER)
    for n, cat, note in rows:
        w.writerow([n, "", cat, note])

xlsx_made = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Mounts - Names"
    hf = PatternFill("solid", fgColor="2F2F2F"); hfont = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(HEADER, 1):
        cell = ws.cell(1, c, h); cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    rus = PatternFill("solid", fgColor="FFF7E6"); opt = PatternFill("solid", fgColor="EEF3FA")
    for r, (n, cat, note) in enumerate(rows, 2):
        ws.cell(r, 1, n)
        rc = ws.cell(r, 2, ""); rc.fill = rus
        ws.cell(r, 3, cat)
        ws.cell(r, 4, note)
        if note == "optional":
            for col in (1, 3, 4):
                ws.cell(r, col).fill = opt
    for i, wd in enumerate([44, 34, 16, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.column_dimensions["A"].alignment = None
    wb.save(os.path.join(OUTDIR, "ui_strings_mount_names.xlsx"))
    xlsx_made = True
except ImportError:
    pass

from collections import Counter
by_cat = Counter(r[1] for r in rows)
print("Mounts-page names: %d unique" % len(rows))
for cat in ["Mount", "Combat Power", "Equip Power", "Insignia Bonus", "Insignia", "Collar"]:
    print("   %-14s %4d%s" % (cat, by_cat.get(cat, 0), "  (optional)" if cat in ("Insignia", "Collar") else ""))
print("   optional (patterned) total: %d" % sum(1 for r in rows if r[2] == "optional"))
print("CSV : %s" % os.path.relpath(csv_path, WEB))
print("XLSX: %s" % ("docs/i18n/ui_strings_mount_names.xlsx  (tab 'Mounts - Names')" if xlsx_made else "skipped"))
