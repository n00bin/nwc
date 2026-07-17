#!/usr/bin/env python3
"""
i18n pilot — export the MOUNTS PAGE UI strings into a translator-friendly sheet
(English filled, Russian blank), to be imported as a NEW TAB in the existing
Google Sheet (File -> Import -> Insert new sheet(s)).

Hand-curated from mounts.html and js/mounts-page.js. Item NAMES, stat values and
numbers are intentionally excluded (item names stay English via translate="no").
Strings already translated on the home tab (I18N_RU) are not repeated.

Output: docs/i18n/ui_strings_mounts.csv  (canonical)
        docs/i18n/ui_strings_mounts.xlsx (sheet/tab named "Mounts")

Run with the interpreter that has openpyxl:
  G:/Python/python.exe scripts/i18n_export_mounts.py
"""
import os, csv, collections

HERE = os.path.dirname(os.path.abspath(__file__))
WEB = os.path.dirname(HERE)
OUTDIR = os.path.join(WEB, "docs", "i18n")
os.makedirs(OUTDIR, exist_ok=True)

# (english, where_it_appears, note)   note: "brand" or "optional" or ""
STRINGS = [
    # --- View tabs ---
    ("Lookup", "Mounts: view tab", ""),
    ("Combat Power Ranking", "Mounts: view tab", ""),
    ("DPS Combat Power Ranking", "Mounts: view tab", ""),
    ("Support Equip Bonuses", "Mounts: view tab", ""),
    ("Collars", "Mounts: view tab", ""),
    ("Insignias", "Mounts: view tab", ""),
    ("Stable Planner", "Mounts: view tab", ""),
    ("Every mount combat power, ranked by overall strength", "Mounts: tooltip on the Combat Power Ranking tab", ""),
    ("Combat powers ranked for single-target DPS only (boss fights)", "Mounts: tooltip on the DPS Ranking tab", ""),
    ("Equip powers ranked for Tank and Healer builds", "Mounts: tooltip on the Support Equip Bonuses tab", ""),

    # --- Search boxes ---
    ("Search mounts...", "Mounts: search box (Lookup)", ""),
    ("Search combat powers...", "Mounts: search box (Combat Power Ranking)", ""),
    ("Search DPS powers...", "Mounts: search box (DPS Ranking)", ""),
    ("Search equip bonuses...", "Mounts: search box (Support Equip Bonuses)", ""),
    ("Search collars...", "Mounts: search box (Collars)", ""),
    ("Search bonuses by name or effect…", "Mounts: search box (Stable Planner bonus picker)", ""),

    # --- Filter dropdowns: 'all' options ---
    ("All Combat Powers", "Mounts: filter dropdown (Lookup)", ""),
    ("All Equip Powers", "Mounts: filter dropdown (Lookup)", ""),
    ("All Insignia Bonuses", "Mounts: filter dropdown (Lookup)", ""),
    ("All Categories", "Mounts: filter dropdown (Collars)", ""),
    ("All Types", "Mounts: filter dropdown (Collars)", ""),
    ("All Names", "Mounts: filter dropdown (Insignias)", ""),
    ("All Qualities", "Mounts: filter dropdown (Insignias)", ""),
    ("All Slot Types", "Mounts: filter dropdown (Insignias)", ""),
    ("Top in General (All)", "Mounts: filter dropdown (DPS Ranking)", ""),
    ("Top Magical", "Mounts: filter dropdown (DPS Ranking)", ""),
    ("Top Physical", "Mounts: filter dropdown (DPS Ranking)", ""),

    # --- Insignia quality tiers (use the official in-game Russian names) ---
    ("Celestial", "Mounts: insignia quality tier (use in-game RU name)", ""),
    ("Mythic", "Mounts: insignia quality tier (use in-game RU name)", ""),
    ("Legendary", "Mounts: insignia quality tier (use in-game RU name)", ""),
    ("Epic", "Mounts: insignia quality tier (use in-game RU name)", ""),
    ("Rare", "Mounts: insignia quality tier (use in-game RU name)", ""),
    ("Uncommon", "Mounts: insignia quality tier (use in-game RU name)", ""),

    # --- Insignia slot types (use the official in-game Russian names) ---
    ("Regal", "Mounts: insignia slot type (use in-game RU name)", ""),
    ("Barbed", "Mounts: insignia slot type (use in-game RU name)", ""),
    ("Crescent", "Mounts: insignia slot type (use in-game RU name)", ""),
    ("Illuminated", "Mounts: insignia slot type (use in-game RU name)", ""),
    ("Enlightened", "Mounts: insignia slot type (use in-game RU name)", ""),

    # --- Lookup checkboxes ---
    ("Preferred Slot", "Mounts: checkbox (Lookup)", ""),
    ("4-Slot Only", "Mounts: checkbox (Lookup)", ""),
    ("Only mounts with a preferred slot: the matching insignia type gets a +20% bonus there",
     "Mounts: tooltip on the Preferred Slot checkbox", ""),
    ("Only mounts with four insignia slots (more room for stat bonuses)",
     "Mounts: tooltip on the 4-Slot Only checkbox", ""),

    # --- Detail panel labels & empty states ---
    ("Pick a mount on the left to see its powers, insignias, and stats.", "Mounts: Lookup, empty detail panel", ""),
    ("Select a mount to view details", "Mounts: Lookup, empty detail panel (short)", ""),
    ("Combat Power", "Mounts: detail section label", ""),
    ("Equip Power", "Mounts: detail section label", ""),
    ("Insignia Slots", "Mounts: detail section label", ""),
    ("Role-specific effects", "Mounts: detail section label", ""),
    ("Stacking Buff", "Mounts: detail label", ""),
    ("Notes", "Mounts: detail section label", ""),
    ("Duration:", "Mounts: detail label (before a value)", ""),
    ("Max stacks:", "Mounts: detail label (before a value)", ""),
    ("Power tier:", "Mounts: detail label (before a value)", ""),
    ("Trigger:", "Mounts: detail label (before a value)", ""),
    ("Source:", "Mounts: detail label (before a value)", ""),

    # --- Table headers (Insignias / Collars / rankings) ---
    ("Name", "Mounts: table column header", ""),
    ("Quality", "Mounts: table column header", ""),
    ("Item Level", "Mounts: table column header", ""),
    ("Combined Rating", "Mounts: table column header", ""),
    ("Stats", "Mounts: table column header", ""),
    ("Slot Types", "Mounts: table column header", ""),
    ("Category", "Mounts: table column header", ""),
    ("Type", "Mounts: table column header", ""),
    ("Bonus", "Mounts: table column header", ""),
    ("Rank", "Mounts: table column header", ""),

    # --- Empty / not-loaded states ---
    ("No mounts match your filters", "Mounts: empty list (Lookup)", ""),
    ("No results match your filters", "Mounts: empty list (rankings)", ""),
    ("No collars match your filters.", "Mounts: empty list (Collars)", ""),
    ("No insignias match your filters.", "Mounts: empty list (Insignias)", ""),
    ("No combat power data", "Mounts: detail, missing data", ""),
    ("No equip power data", "Mounts: detail, missing data", ""),
    ("No insignia slots", "Mounts: detail, missing data", ""),
    ("No compatible insignia bonuses", "Mounts: planner, none found", ""),
    ("No mount in the database can host this bonus.", "Mounts: planner, none found", ""),
    ("Collar data is not available.", "Mounts: Collars, data missing", ""),
    ("Insignia data not loaded.", "Mounts: Insignias, data missing", ""),

    # --- Stable Planner: controls & labels ---
    ("+ Add Loadout", "Mounts: Stable Planner button", ""),
    ("Clear All", "Mounts: Stable Planner button", ""),
    ("Saved in this browser only.", "Mounts: Stable Planner, note next to buttons", ""),
    ("Delete", "Mounts: Stable Planner, delete a loadout", ""),
    ("Pick an Insignia Bonus", "Mounts: Stable Planner, picker heading", ""),
    ("Selected Insignia Bonus", "Mounts: Stable Planner, picker heading", ""),
    ("My insignias for this loadout:", "Mounts: Stable Planner, section label", ""),
    ("What can I build with my insignias?", "Mounts: Stable Planner button", ""),
    ("What can I build?", "Mounts: Stable Planner button (short)", ""),
    ("What you can build now", "Mounts: Stable Planner, results heading", ""),
    ("Insignia Sharing Plan", "Mounts: Stable Planner, results heading", ""),
    ("Excluded mounts", "Mounts: Stable Planner, excluded list heading", ""),
    ("Restore all", "Mounts: Stable Planner, restore excluded mounts", ""),
    ("reused", "Mounts: Stable Planner badge", ""),
    ("shared", "Mounts: Stable Planner badge", ""),
    ("single use — no sharing possible", "Mounts: Stable Planner badge", ""),
    ("all instances in one loadout — already optimal", "Mounts: Stable Planner badge", ""),

    # --- Roles (may be left English if that reads better in RU) ---
    ("DPS", "Mounts: build role", ""),
    ("Tank", "Mounts: build role", ""),
    ("Healer", "Mounts: build role", ""),

    # --- Stable Planner how-to (LONG help text — optional, can be left for later) ---
    ("How to use the Stable Planner", "Mounts: Stable Planner help title", ""),
    ("Click \"+ Add Loadout\" for each loadout you actually run in-game (e.g., one for healer, one for tank, one per DPS spec).",
     "Mounts: Stable Planner how-to, step 1", "optional"),
    ("Name it, pick a role (DPS / Tank / Healer), then add up to 5 insignia bonuses you want active in that loadout. The same bonus can be added more than once if you want it on multiple mounts.",
     "Mounts: Stable Planner how-to, step 2", "optional"),
    ("Fill in \"My insignias for this loadout\" on each loadout with how many of each type you own — just five numbers, no sorting needed. Each loadout has its own counts, so you can plan loadouts with different insignia sets.",
     "Mounts: Stable Planner how-to, step 3", "optional"),
    ("Click \"What can I build with my insignias?\" on a loadout. It shows which of that loadout's bonuses you can run at the same time with what you own (no insignia doing double duty), and exactly how many more of each type you'd need for the rest.",
     "Mounts: Stable Planner how-to, step 4", "optional"),
    ("Check the Insignia Sharing Plan below to go further. For each bonus, it shows the minimum number of distinct mounts you need across all your loadouts — and exactly which mounts to pick — so the same insignia set covers every loadout that wants the bonus. Each row has a \"Save N upgrades\" badge.",
     "Mounts: Stable Planner how-to, step 5", "optional"),
    ("★ markers mean a mount has a preferred slot that can be filled with its preferred type (+20% IL & stats on that insignia). Those mounts are listed first.",
     "Mounts: Stable Planner how-to, step 6", "optional"),
    ("Don't own a recommended mount? Click the × on its tile. It moves to the \"Excluded mounts\" bar and the next-best candidate takes its place. Click ↻ to restore.",
     "Mounts: Stable Planner how-to, step 7", "optional"),
    ("Everything saves automatically in this browser only — refresh, come back tomorrow, your loadouts are still there. Use \"Clear All\" to reset.",
     "Mounts: Stable Planner how-to, step 8", "optional"),
]

# Dedupe by English string, and drop anything already translated on the home tab.
HOME_KEYS = set()
home_js = os.path.join(WEB, "data", "i18n-ru.js")
if os.path.exists(home_js):
    import re
    txt = open(home_js, encoding="utf-8").read()
    # keys are the quoted strings on the left of a ':' — grab the first quoted
    # literal of each mapping line; good enough to avoid obvious repeats.
    for m in re.finditer(r'"((?:[^"\\]|\\.)*)"\s*:', txt):
        HOME_KEYS.add(m.group(1))

seen = collections.OrderedDict()
skipped_dupe = []
for eng, ctx, note in STRINGS:
    if eng in HOME_KEYS:
        skipped_dupe.append(eng)
        continue
    if eng in seen:
        continue
    seen[eng] = (ctx, note)

rows = [(e, c, n) for e, (c, n) in seen.items()]
HEADER = ["English (do not edit)", "Russian (fill this in)", "Where it appears", "Note"]

csv_path = os.path.join(OUTDIR, "ui_strings_mounts.csv")
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(HEADER)
    for e, c, n in rows:
        w.writerow([e, "", c, n])

xlsx_made = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Mounts"
    hf = PatternFill("solid", fgColor="2F2F2F"); hfont = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(HEADER, 1):
        cell = ws.cell(1, c, h); cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    rus = PatternFill("solid", fgColor="FFF7E6"); opt = PatternFill("solid", fgColor="EEF3FA")
    for r, (e, c, n) in enumerate(rows, 2):
        ws.cell(r, 1, e)
        rc = ws.cell(r, 2, ""); rc.fill = rus
        ws.cell(r, 3, c)
        nc = ws.cell(r, 4, n)
        if n == "optional":
            for col in range(1, 5):
                ws.cell(r, col).fill = opt
            rc.fill = rus
    for i, wd in enumerate([56, 34, 46, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    for r in range(1, len(rows) + 2):
        for c in (1, 3):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(os.path.join(OUTDIR, "ui_strings_mounts.xlsx"))
    xlsx_made = True
except ImportError:
    pass

print("Mounts UI strings (unique): %d" % len(rows))
print("  of which optional (planner how-to): %d" % sum(1 for _, _, n in rows if n == "optional"))
print("  skipped (already on home tab): %d %s" % (len(skipped_dupe), skipped_dupe or ""))
print("CSV : %s" % os.path.relpath(csv_path, WEB))
print("XLSX: %s" % ("docs/i18n/ui_strings_mounts.xlsx" if xlsx_made else "skipped (no openpyxl)"))
